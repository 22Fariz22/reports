# -*- coding: utf-8 -*-
import pandas as pd
from datetime import *
import email
import smtplib
import os
import openpyxl
from openpyxl import load_workbook
import cx_Oracle
import shutil
import sys

if sys.platform.startswith("darwin"):
    instant_client_dir = os.environ.get("HOME")+"/Downloads/instantclient_19_8"
if instant_client_dir is not None:
    cx_Oracle.init_oracle_client(lib_dir=instant_client_dir)

flag_listdir = False
flag_names_files_defined = False
flag_trunc = False
flag_sql_oracle = False
flag_delta = False
flag_check = False
flag_begin_end = False
flag_new_files = False
flag_move_reports = False

sum_sales_reports = 0
sum_credit_report = 0
filename_own = ''
filename_franch = ''

source_dir = '/Users/farizrustamov/GOOD_COMPANY/Отчет/test_for_partner'

def write(text):
    with open("log.txt", "a") as file:
        file.write(f'{text}')

x = os.listdir(source_dir)
for i in x:
    if i[0:5].lower() == 'отчет':
        flag_listdir = True
        break

salon_col = 0
dt = datetime.now()
write(f' \n')
write(f'Начало скрипта: {dt}.\n')
if flag_listdir == True:
    for i in x:
        if i[0:5] == 'Отчет':
            name_file = i
            name_file_with_date = name_file
            new_file_name_2 = '2 ' + name_file
            wb = load_workbook(filename=f'{source_dir}/{i}', data_only=True)
            wb.active = 0
            sheet = wb.active

            wb.active = 1
            sheet2 = wb.active

            all_data = []
            first_data = []
            last_data = []

            data = [sheet2.cell(row=4, column=i).value for i in
                    range(1, 500)]
            for i in data:
                if i != None:
                    all_data.append(i)
            first_data = [all_data[0]]
            last_data = [all_data[-1]]

            yyyy, mm, dd = first_data[0].split('/')
            first_data[0] = '{}-{}-{}'.format(yyyy,mm,dd[0:2])

            first_data_1 = datetime.strptime(first_data[0], '%Y-%m-%d')
            sheet['B2'].value = first_data_1

            yyyy, mm, dd = last_data[0].split('/')

            last_data[0] = '{}-{}-{}'.format(yyyy,mm,dd[0:2])
            last_data_1 = datetime.strptime(last_data[0], '%Y-%m-%d')
            sheet['C2'].value = last_data_1
            data2 = [sheet.cell(row=8, column=i).value for i in range(1, 8)]
            maxrow = sheet.max_row

            for e,i in enumerate(data2):
                if 'Салон' in str(i):
                    if sheet.cell(maxrow-3,1).value != 'Местонахождение':
                        sheet.cell(maxrow-3,1).value = 'Местонахождение'
                    salon_col=e
                    filename_own = name_file
                    break
            else:
                filename_franch = name_file
                wb.close()
    flag_names_files_defined = True

write(f'Файл Собственные: {filename_own}\n' )
write(f'Файл Франчи: {filename_franch}\n' )
write(f'first_data: {first_data_1}\n')
write(f'last_data: {last_data_1}\n')

def report1_daily(file, sheet):
    global salon_col
    df_input = pd.read_excel(file,sheet_name=sheet, skiprows=3)
    columns = df_input.columns
    dates = []
    for col in columns:
        if col.startswith('Unnamed'):
            continue
        else:
            dates.append(col)

    num_column_data = 0
    for e,col in enumerate(df_input.columns[0:]):
        if str(col)[:3] == '202':
            num_column_data = e
            break
    for col in df_input.columns[num_column_data:]:
        if col.startswith('Unnamed'):
            df_input.rename(columns={col: last_date}, inplace=True)
        else:
            last_date = col
    banks = list(set(df_input.loc[0]))[1:]
    df_input.rename(columns={'Unnamed: 0': 'Дирекция', 'Unnamed: 1': 'Город', 'Unnamed: 2':'Магазин', 'Unnamed: 3': 'Тип продажи'}, inplace=True)
    df_input.fillna('None', inplace=True)
    channel = "РБТ"
    output_df2 = []
    last_direction = df_input.loc[1, 'Дирекция']
    last_town = df_input.loc[1, 'Город']

    for i in range(1, df_input.shape[0]):
        if i == df_input.shape[0]:
            break
        if i != df_input.shape[0]:
                last_direction = df_input.loc[i, 'Дирекция']
        if i != df_input.shape[0]:
                last_town = df_input.loc[i, 'Город']

        for j in range(num_column_data,len(df_input.columns[num_column_data:]) + num_column_data): # проверить!!
            if df_input.iloc[i, j] == 0:
                continue
            output_list = []
            output_list.append('d') #d_type
            output_list.append(channel) #channel
            output_list.append(df_input.iloc[i, salon_col]) #sellerplace_group
            output_list.append(datetime.strptime(df_input.columns[j].strip(), '%Y/%m/%d')) #date_
            output_list.append(df_input.iloc[0, j]) #competitor
            output_list.append(df_input.iloc[i, j]) #credit_amt
            output_list.append(date.today()) #date_update
            output_df2.append(output_list)

    out_to_excel = pd.DataFrame(data=output_df2, columns=['D_TYPE'.lower(), 'CHANNEL'.lower(),
                                                          'SELLERPLACE_GROUP'.lower(), 'DATE_'.lower(),
                                                          'COMPETITOR'.lower(), 'CREDIT_AMT'.lower(),
                                                          'DATE_UPDATE'.lower()])

    return list(out_to_excel.values.tolist()), out_to_excel.columns


def report1_period(file, sheet, period_start, period_end):
    global salon_col
    df_input = pd.read_excel(file,sheet_name=sheet, skiprows=2)
    columns_list = []
    df_input.rename(columns={'Показатели\nas values':'Дирекция', 'Unnamed: 1':'Город', 'Unnamed: 2':'Магазин', 'Unnamed: 3':'Тип продажи'}, inplace=True)
    columns_lower = [column.lower() for column in list(df_input.columns)]
    for column in df_input.columns[columns_lower.index('итого'):columns_lower.index('итого') + 3]:
        columns_list.append(df_input.loc[0, column])
        df_input.rename(columns={column: df_input.loc[0, column]}, inplace=True)

    df_input.fillna('None', inplace=True)
#    ----------------------------------------------------------
    channel = "GOODCOMPANY"
    last_direction = df_input.loc[1, 'Дирекция']
    last_town = df_input.loc[1, 'Город']
    data = []

    for i in range(1, df_input.shape[0]):
        if df_input.loc[i, columns_list[0]] == 0:
                continue

        if i == df_input.shape[0]-4:  #'Местонахождение':
            break

        if i != df_input.shape[0]-4:
                last_direction = df_input.loc[i, 'Дирекция']

        if i != df_input.shape[0]-4:
                last_town = df_input.loc[i, 'Город']
        need_data = []
        need_data.append(df_input.iloc[i, salon_col]) #change магазин to salon
        need_data.append(df_input.loc[i, columns_list[0]])
        need_data.append(df_input.loc[i, columns_list[2]])
        need_data.append(channel)
        need_data.append('p')
        need_data.append(period_start)
        need_data.append(period_end)
        need_data.append(date.today())
        data.append(need_data)

    new_to_excel_DF = pd.DataFrame(data=data, columns=['sellerplace_group', 'sales_amt', 'sh_ko', 'channel', 'd_type', 'date_','date_to', 'date_update'])
    return list(new_to_excel_DF.values.tolist()), new_to_excel_DF.columns

def report2_daily(file, sheet):
    df_input = pd.read_excel(file, sheet_name=sheet, skiprows=3)
    columns = df_input.columns
    dates = []
    for col in columns:
        if col.startswith('Unnamed'):
            continue
        else:
            dates.append(col)
    banks = list(set(df_input.loc[0]))[1:]
    for col in df_input.columns[1:]:
        if col.startswith('Unnamed'):
            df_input.rename(columns={col: last_date}, inplace=True)
        else:
            last_date = col
    df_input.rename(columns={'Unnamed: 0': 'Дирекция'}, inplace=True)
    channel = "GOODCOMPANY ФР"
    output_df2 = []
    last_direction = df_input.loc[1, 'Дирекция']

    for i in range(1, df_input.shape[0]):
        if i == df_input.shape[0]:
            break
        if i == df_input.shape[0]:
            last_direction = df_input.loc[i, 'Дирекция']
        for j in range(1, len(df_input.columns[1:]) + 1):
            if df_input.iloc[i, j] == 0:
                continue
            output_list = []
            output_list.append('d')
            output_list.append(channel)
            output_list.append(df_input.loc[i, 'Дирекция'])
            output_list.append(datetime.strptime(df_input.columns[j].strip(), '%Y/%m/%d'))
            output_list.append(df_input.iloc[0, j])
            output_list.append(df_input.iloc[i, j])
            output_list.append(date.today())
            output_df2.append(output_list)
    out_to_excel = pd.DataFrame(data=output_df2,
                                columns=['d_type', 'channel', 'sellerplace_group', 'date_', 'competitor', 'credit_amt',
                                         'date_update'])
    return list(out_to_excel.values.tolist()), out_to_excel.columns

def report2_period(file, sheet, period_start, period_end):
    df_input = pd.read_excel(file,sheet_name=sheet, skiprows=2) # Вставляем Excel файл с данными
    df_input.rename(columns={'Показатели\nas values':'Дирекция'}, inplace=True)
    columns_list = []
    columns_lower = [column.lower() for column in list(df_input.columns)]


    for column in df_input.columns[columns_lower.index('итого'):columns_lower.index('итого') + 3]:
        columns_list.append(df_input.loc[0, column])
        df_input.rename(columns={column: df_input.loc[0, column]}, inplace=True)

    channel = "GOODCOMPANY ФР"
    last_direction = df_input.loc[1, 'Дирекция']
    output_df2 = []

    for i in range(1, df_input.shape[0]):
        if df_input.loc[i, columns_list[0]] == 0:
            continue

        if i == df_input.shape[0] - 1:
            break
        if i != df_input.shape[0] - 1:
                last_direction = df_input.loc[i, 'Дирекция']
        output_list = []
        output_list.append(last_direction)
        output_list.append(df_input.loc[i, columns_list[0]]) # sales_amt
        output_list.append(df_input.loc[i, columns_list[2]])
        output_list.append(channel)
        output_list.append('p')
        output_list.append(period_start)
        output_list.append(period_end)
        output_list.append(date.today())
        output_df2.append(output_list)
    new_to_excel_DF = pd.DataFrame(data=output_df2, columns=['sellerplace_group', 'sales_amt', 'sh_ko', 'channel', 'd_type', 'date_','date_to', 'date_update'])
    return list(new_to_excel_DF.values.tolist()), new_to_excel_DF.columns

def actual_columns(df_input):
    k = 0
    i = 0
    for col in df_input.columns[4:]:
        if col.startswith("Unnamed"):
            k += 1
        else:
            k = 0
        if k >= 2:
            break
        i += 1
    return i + 4

columns_with_type = [('channel', 'varchar'), ('is_online', 'number'), ('sellerplace_group', 'varchar'), ('d_type', 'varchar'), ('date_', 'date'), ('date_to', 'date'), ('competitor', 'varchar'), ('credit_amt', 'number'), ('sh_ko', 'number'), ('sales_amt', 'number'), ('sh_comp', 'float'), ('date_update', 'date')]


def make_connection(db_name):

    if db_name == 'dwhcor':
        user = 'FRUSTAMOV'
        passw = 'sdf&8sd'
    else:
        user = 'error'
        passw = 'error'
    if db_name == 'dwhcor':
        ip = '123.123.123.123'
        port = 1111
        service_name = 'DWHCOR.primary.prod.db.goodcompany'
        dsn = cx_Oracle.makedsn(ip, port, service_name=service_name)
    else:
        dsn = 'error'
    conn = cx_Oracle.connect(user, passw, dsn, encoding='UTF-8')
    return conn


def to_sql_(data, columns_name, db_name='dwhcor'):
    try:
        conn = make_connection(db_name)
        write(f'Make_connection(db_name) and db_name: {db_name}.\n')
    except Exception as err:
        write(f'Error while creating connection: {err}.\n' )
        write(f'Except in try conn = make_connection(db_name). db_name: {db_name}.\n')
    else:
        try:
            cursor = conn.cursor()
            all_column_names = [col[0] for col in columns_with_type]
            columns_string = ''
            for column in columns_name:
                columns_string += column
                if column != columns_name[-1]:
                    columns_string += ', '
            for row in data:
                row_string = ''
                for element in row:
                    if (columns_with_type[all_column_names.index(columns_name[row.index(element)])][1] == 'varchar'):
                        row_string += f"'{element}'"
                    elif ((columns_with_type[all_column_names.index(columns_name[row.index(element)])][
                               1] == 'number') or (
                                  columns_with_type[all_column_names.index(columns_name[row.index(element)])][
                                      1] == 'float')):
                        row_string += str(element)
                    elif columns_with_type[all_column_names.index(columns_name[row.index(element)])][1] == 'date':
                        if len(str(element)) > 10:
                            row_string += "to_date(" + f"'{element}'" + ", 'yyyy-mm-dd hh24:mi:ss')"
                        else:
                            row_string += "to_date(" + f"'{element}'" + ", 'yyyy-mm-dd')"
                    if element != row[-1]:
                        row_string += ', '
                sql_insert = f"""
                INSERT INTO FRUSTAMOV.goodcompany_PARTNERS_DATA({columns_string}
                )
                VALUES ({row_string})
                """
                cursor.execute(sql_insert)
        except Exception as err:
            write(f'Error while inserting the data: {err}.\n' )
            write(f'{type(columns_string)}.\n')
            write(f'Columns_string: {columns_string}. \n')
            write(f'{type(row_string)}.\n')
            write(f'Row_string: {row_string}.\n')

        else:
            write(f'Inserting completed.\n')
            conn.commit()
    finally:
        cursor.close()
        conn.close()

def trunc_table():
    global flag_trunc
    write('Start truncating oracle table.\n')
    if flag_names_files_defined == True:
        db_name = 'dwhcor'
        try:
            conn = make_connection(db_name)
            write(f'Try conn = make_connection(db_name) and db_name: {db_name}.\n')
        except Exception as err:
            write(f'Error while creating connection: {err}.\n')
            write(f'Here except in try onn = make_connection(db_name). db_name: {db_name}.\n')
        else:
            try:
                cursor = conn.cursor()
                sql_truncate = f"""
                                TRUNCATE TABLE vdd_report.goodcompany_partners_data_in
                                """
                cursor.execute(sql_truncate)
            except Exception as err:
                write(f'Error while: {err}\n.')
            else:
                flag_trunc = True
                write(f'Truncating completed, flag_trunc = True.\n')
                conn.commit()
        finally:
            cursor.close()
            conn.close()

def check():
    global flag_check
    global flag_delta
    global flag_sql_oracle
    if flag_sql_oracle == True:
        db_name = 'dwhcor'
        try:
            conn = make_connection(db_name)
            write(f'Here try conn = make_connection(db_name) and db_name: {db_name}.\n')
        except Exception as err:
            write(f'Error while creating connection: {err}.')
            write(f'Except in try onn = make_connection(db_name). db_name:  {db_name}.\n')
        else:
            try:
                cursor = conn.cursor()
                sql_check_rows = f"""
                                     SELECT count(*) FROM FRustamov.goodcompany_Partners_Data
                                     """
                cursor.execute(sql_check_rows)
                res_check_rows_sql = cursor.fetchall()
                cursor = conn.cursor()
                sql_check_sales_amt = f"""
                                          SELECT sum(SALES_AMT) FROM FRustamov.goodcompany_Partners_Data
                                          """
                cursor.execute(sql_check_sales_amt)
                res_sales_oracle = cursor.fetchall()
                cursor = conn.cursor()
                sql_check_cred_amt = f"""
                                         SELECT sum(CREDIT_AMT) FROM FRustamov.goodcompany_Partners_Data
                                         """
                cursor.execute(sql_check_cred_amt)
                res_cred_oracle = cursor.fetchall()
                cursor = conn.cursor()
                check_delta = f"""
                                            select trunc(t.date_,'mm') as mnth, t.channel,t.sellerplace_group,sum(t.credit_amt) as camt,
                                            sum(t.sales_amt) as samt,
                                            sum(sh_ko*sales_amt) as sh_amt,
                                            abs(nvl(sum(t.credit_amt),0)-nvl(sum(sh_ko*sales_amt),0)) as delta
                                            from frustamov.rbt_raw_penetr_partners_data t
                                            where channel='РБТ' and trunc(t.date_,'mm')=to_date('01.06.2021','dd.mm.yyyy') 
                                            group by trunc(t.date_,'mm'),t.channel,t.sellerplace_group
                                            having abs(nvl(sum(t.credit_amt),0)-nvl(sum(sh_ko*sales_amt),0))>1
                                                  """

                cursor.execute(check_delta)
                delta = cursor.fetchall()
            except Exception as err:
                write(f'Error while: {err}.\n')
            else:
                write(f'Fetching completed.\n')
                conn.commit()
        finally:
            cursor.close()
            conn.close()

        sales_amt_in_reports = 0
        cred_amt_in_reports = 0

        x = os.listdir(source_dir)
        for i in x:
            if i[0:5] == 'Отчет':
                name_file = i
                wb = load_workbook(filename=f'{source_dir}/{i}',
                                   data_only=True)
                sheet = wb.active
                for i in range(1, 500):
                    if sheet.cell(row=3, column=i).value == None:
                        continue
                    if sheet.cell(row=3, column=i).value.lower() == 'итого':
                        for sale in range(5, 2000):
                            if sheet.cell(row=sale, column=i).value == None:
                                sales_amt_in_reports += sheet.cell(row=sale - 1, column=i).value
                                break
                        for credit in range(5, 2000):
                            if sheet.cell(row=credit, column=i + 1).value == None:
                                cred_amt_in_reports += sheet.cell(row=credit - 1, column=i + 1).value
                                break
        write(f'amount sales in reports: {sales_amt_in_reports}.\n')
        write(f'amount credit in reports: {cred_amt_in_reports}.\n')
        if len(delta)>0:
            write(f'Delta is: {delta}.\n')
            write(f'Amount of deltas: {len(delta)}.\n')
            flag_delta = True
        else:
            write(f'No deltas.\n')
        write(f'Amount of sales in oracle_bd: {res_sales_oracle[0][0]}. \n')
        write(f'Amount of credits in oracle_bd: {res_cred_oracle[0][0]}. \n')
        write(f'Rows in oracle_bd: {res_check_rows_sql[0][0]}.\n' )
        if sales_amt_in_reports == res_sales_oracle[0][0] and cred_amt_in_reports == res_cred_oracle[0][0] and flag_delta == False:
            flag_check = True
            write(f'Data from reports and from oracle_bd coincide, there is no delta. flag_check = True.\n')
        else:
            print(f'Data does not match. flag_delta = {flag_delta}, flag_check={flag_check}.')

def begin_end():
    global flag_check
    global flag_begin_end
    write('Start procedure.\n')
    db_name = 'dwhcor'
    if flag_check == True:
        try:
            conn = make_connection(db_name)
            write(f'Try conn = make_connection(db_name) and db_name: {db_name}.\n')
        except Exception as err:
            write(f'Error while creating connection: {err}.\n')
            write(f'Here except in try conn = make_connection(db_name). db_name: {db_name}.\n')
        else:
            try:
                cursor = conn.cursor()
                sql_begin_end = f"""
                                BEGIN
                                    rustamov.up_partner_report2()
                                END
                                """
                cursor.execute(sql_begin_end)
            except Exception as err:
                write(f'Error while {err}.\n')
            else:
                conn.commit()
                flag_begin_end = True
                write(f'Procedura Begin_End completed. flag_begin_end = True.\n')
        finally:
            cursor.close()
            conn.close()

trunc_table()

if flag_trunc == True:
    wb = openpyxl.load_workbook(filename = f'{source_dir}/{filename_own}', data_only=True)
    write(f'Start report1_period().\n')
    data, columns_name = report1_period(f'{source_dir}/{filename_own}', sheet=0, period_start=first_data_1, period_end=last_data_1)
    write(f'Start to_sql_().\n')
    to_sql_(data, columns_name)
    write(f'Start report1_daily().\n')
    data, columns_name = report1_daily(f'{source_dir}/{filename_own}', sheet=1)
    write(f'Start to_sql_().\n')
    to_sql_(data, columns_name)
    write(f'End to_sql file with Own.\n')

    write(f'Start to_sql file with franch.\n')
    wb = openpyxl.load_workbook(filename=f'{source_dir}/{filename_franch}', data_only=True)
    write(f'Start report2_period().\n')
    data, columns_name = report2_period(f'{source_dir}/{filename_franch}', 0, first_data_1, last_data_1)
    write(f'Start to_sql_().\n')
    to_sql_(data, columns_name)
    write(f'Start report2_daily().\n')
    data, columns_name = report2_daily(f'{source_dir}/{filename_franch}', 1)
    write(f'Start to_sql_().\n')
    to_sql_(data, columns_name)
    flag_sql_oracle = True
    write(f'End to_sql file with Franch.\n')
    write(f'Flag_sql_oracle = True.\n')
    wb.close()

check()

# begin_end()

if flag_begin_end==True:
    target_dir = f'{source_dir}/archive'
    now = datetime.now()
    current_time = now.strftime("%Y-%m-%d %H-%M-%S")
    for file in (filename_own, filename_franch):
        try:
            shutil.move(os.path.join(source_dir, file), target_dir)
            os.rename(f'{target_dir}/{file}', f'{target_dir}/{file[0:-5]} {current_time}{file[-5:]}')
            write(f'File {file} is moved and renamed.\n')
            flag_move_reports = True
        except (shutil.Error, OSError):
            write(f'Error: {shutil.Error, OSError}.\n')
            flag_move_reports = False
            break


if flag_check==True:
    date_today = datetime.date.today()
    target_dir = f'{source_dir}/{date_today}'
    try:
        os.mkdir(target_dir)
    except OSError:
        print ("Creation of the directory %s failed" % target_dir)
    else:
        print ("Successfully created the directory %s " % target_dir)

    file_names = os.listdir(source_dir)
    for file in file_names:
        if file[0:5].lower() == 'отчет':
            try:
                shutil.move(os.path.join(source_dir, file), target_dir)
            except (shutil.Error, OSError):
                print('error: ',shutil.Error, OSError)
        if file[0:3].lower() == 'new' or file[0:3].lower() == '1 о' or file[0:3].lower() == '2 о':
            try:
                os.remove(f'{source_dir}/{file}')
            except OSError as e:
                print("Ошибка: %s : %s" % (file, e.strerror))
