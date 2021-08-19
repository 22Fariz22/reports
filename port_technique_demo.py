# -*- coding: utf-8 -*-
import datetime
from datetime import *
import os
import openpyxl
from openpyxl import load_workbook
import cx_Oracle
import shutil
import sys

if sys.platform.startswith("darwin"):
    instant_client_dir = os.environ.get("HOME")+"/Downloads/instantclient_19_8"
if instant_client_dir is not None:
    cx_Oracle.init_oracle_client(lib_dir=instant_client_dir)

table = 'FRustamov.Partners_Data'
source_dir = '/Users/farizrustamov/good_company/Отчет **/Портативная техника'
target_dir = '/Users/farizrustamov/good_company/Отчет **/Портативная техника/archive'
channels = ['Сеть:iFort','Сеть:LCCompany','Сеть:Nokia','Сеть:Huawei Centre','СЦ iFort','СЦ Huawey']
flag_listdir = False
flag_sql_list = False
flag_end_insert_sql = False
flag_move_reports = False
data = []
files = []
itogo = 0
word = ''
sql_list = []

x = os.listdir(source_dir)
for i in x:
    if i[-5:] == '.xlsx':
        files.append(i)
        flag_listdir = True

if flag_listdir == True:
    for i in files:
        name_file = i
        wb = load_workbook(filename=f'{source_dir}/{name_file}', data_only=True)
        wb.active = 0
        sheet = wb.active

        for i in range(4, 1000,2):
            if sheet.cell(row=1, column=i).value != 'Итого':
                for j in range(3, 1000):
                    if sheet.cell(row=j, column=1).value != 'Итого':
                        if sheet.cell(row=j, column=1).value in channels:
                            word = sheet.cell(row=j, column=1).value
                        sql_list.append([word,sheet.cell(row=j, column=1).value,sheet.cell(row=1, column=i).value,sheet.cell(j,i).value,sheet.cell(j,i+1).value])
                    else:
                        break
            else:
                itogo = i
                break
wb.close()

for i,e in enumerate(sql_list):
    if e[0] == e[1]:
        del sql_list[i]
for i in sql_list:
    dd, mm, yyyy = i[2].split('.')
    i[2] = f'{yyyy}-{mm}-{dd}'
    if i[4] == None:
        i[4]= 0
    if i[3] == None:
        i[3]= 0

flag_sql_list = True

def make_connection(db_name):
    if db_name == 'dwhcor':
        user = 'FRUSTAMOV'
        passw = 'sdff4657hDF'
    else:
        user = 'error'
        passw = 'error'
    if db_name == 'dwhcor':
        ip = '192.123.112.11'
        port = 1521
        service_name = 'DWHCOR.primary.prod.db.goodcompany'
        dsn = cx_Oracle.makedsn(ip, port, service_name=service_name)
    else:
        dsn = 'error'
    conn = cx_Oracle.connect(user, passw, dsn, encoding='UTF-8')
    return conn

def to_sql_(data,db_name='dwhcor'):
    global flag_end_insert_sql
    try:
        conn = make_connection(db_name)
        print('here try conn = make_connection(db_name) and db_name - ', db_name)
    except Exception as err:
        print("Error while creating connection", err)
        print('Here except in try onn = make_connection(db_name). db_name - ', db_name)
    else:
        try:
            cursor = conn.cursor()
            for row in data:
                sql_insert = f"""
                INSERT INTO FRUSTAMOV.GOOD_PARTNERS_DATA
                (channel, sellerplace_group,  date_, sales_amt, credit_amt)
                VALUES
                 ('{row[0]}', '{row[1]}', TO_DATE('{row[2]}' , 'yyyy-mm-dd'), '{row[3]}', '{row[4]}')
                """
                cursor.execute(sql_insert)
        except Exception as err:
            print("Error while inserting the data", err)
            print(type(row))
            print("row - ", row)

        else:
            flag_end_insert_sql = True
            print("Inserting completed")
            conn.commit()
    finally:
        cursor.close()
        conn.close()

if flag_sql_list == True:
    to_sql_(sql_list)


if flag_end_insert_sql==True:
    target_dir = f'{source_dir}/archive'
    now = datetime.now()
    current_time = now.strftime("%Y-%m-%d %H-%M-%S")
    for file in (files):
        try:
            shutil.move(os.path.join(source_dir, file), target_dir)
            os.rename(f'{target_dir}/{file}', f'{target_dir}/{file[0:-5]} {current_time}{file[-5:]}')
            print(f'File {file} is moved and renamed.')
            flag_move_reports = True
        except (shutil.Error, OSError):
            print(f'Error: {shutil.Error, OSError}.')
            flag_move_reports = False
            break
